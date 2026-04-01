"""Importación de datos desde archivos Excel/CSV a SQLite.

Soporta importación de ventas (sales), leads, y archivos mixtos.
Incluye validación de cabeceras, normalización de datos, detección de
duplicados y registro de metadata de importación.
"""

import csv
import hashlib
import os
import unicodedata
import logging
from logging.handlers import RotatingFileHandler
from datetime import date, datetime
#from pathlib import Path
from .db import DB_NAME, get_connection, BASE_PATH

# Lazy import de openpyxl (solo cargar cuando se use import de Excel)
# Esto evita problemas al compilar con PyInstaller
_openpyxl = None

def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        from openpyxl import load_workbook
        _openpyxl = load_workbook
    return _openpyxl

_LOGGER = None
MAX_IMPORT_FILE_SIZE_MB = 50  # límite razonable desktop
MAX_IMPORT_FILE_SIZE_BYTES = MAX_IMPORT_FILE_SIZE_MB * 1024 * 1024

# Nombres lógicos obligatorios para guardar en BD
REQUIRED_LOGICAL_COLS = {"client_name", "client_id", "product_type", "price"}

# Normalización de fuentes de marketing (agnóstico a plataforma)
# ============================================================
# Durante la importación de leads, el campo "source" se normaliza
# para unificar valores de distintas plataformas (Meta, Google, LinkedIn, TikTok, etc.)
# Esto permite:
# - Agrupar correctamente leads y gastos por plataforma
# - Calcular métricas consistentes de ROI / CAC / CPL
# - Evitar fragmentación de datos
# 
# Esta normalización DEBE usarse en TODOS los importadores de lead-related data:
# - import_leads_to_db()
# - futuros: import_marketing_spend_to_db()
# - futuros: import_campaigns_to_db()
SOURCE_NORMALIZATION = {
    # Meta ecosystem
    "facebook": "meta_ads",
    "fb": "meta_ads",
    "meta": "meta_ads",
    "instagram": "meta_ads",
    "ig": "meta_ads",
    "messenger": "meta_ads",
    "audience_network": "meta_ads",
    
    # Google ecosystem
    "google": "google_ads",
    "google_ads": "google_ads",
    "adwords": "google_ads",
    "gads": "google_ads",
    
    # LinkedIn
    "linkedin": "linkedin_ads",
    "linkedin_ads": "linkedin_ads",
    
    # TikTok
    "tiktok": "tiktok_ads",
    "tiktok_ads": "tiktok_ads",
    
    # Manual imports
    "csv": "manual_import",
    "manual": "manual_import",
}

# Nombres lógicos obligatorios para leads
REQUIRED_LEADS_COLS = {"created_at", "source"}

# ============================================================
#  MARKETING SPEND — Aliases y columnas requeridas
# ============================================================
REQUIRED_MARKETING_SPEND_COLS = {"source", "amount"}
OPTIONAL_MARKETING_SPEND_COLS = {"date"}

MARKETING_SPEND_HEADER_ALIASES = {
    "source": [
        "source", "platform", "channel", "fuente", "plataforma", "canal",
    ],
    "amount": [
        "amount", "spend", "investment", "importe", "gasto",
        "inversion", "inversión", "monto", "cost", "coste",
    ],
    "date": [
        "date", "fecha", "period", "periodo", "month", "mes",
    ],
}
# Nombres lógicos opcionales para leads
OPTIONAL_LEADS_COLS = {"external_id", "status", "meeting_date", "client_id"}

# Posibles aliases para leads (agnóstico a la plataforma)
LEADS_HEADER_ALIASES = {
    "external_id": [
        "leadgen_id",
        "gclid",
        "lead_id",
        "external_id",
        "campaign_id",
    ],
    "created_at": [
        "created_time",
        "conversion_time",
        "date",
        "created_at",
        "datetime",
    ],
    "source": [
        "source",
        "platform",
        "origin",
        "channel",
    ],
    "status": [
        "status",
        "state",
        "lead_status",
        "campaign_status",
    ],
    "meeting_date": [
        "meeting_date",
        "appointment_date",
        "scheduled_date",
    ],
    "client_id": [
        "client_id",
        "cliente_id",
        "identificador",
    ],
}

# Posibles alias que pueden aparecer en CSV o Excel
HEADER_ALIASES = {
    "cliente": "client_name",
    "nombre cliente": "client_name",
    "nombre": "client_name",
    "client_name": "client_name",  # ← AÑADE ESTA LÍNEA
    "identificador del cliente": "client_id",
    "cliente id": "client_id",
    "id cliente": "client_id",
    "identificador": "client_id",
    "client_id": "client_id",  # ← AÑADE ESTA LÍNEA
    "tipo de producto": "product_type",
    "producto": "product_type",
    "tipo producto": "product_type",
    "tipo": "product_type",
    "product_type": "product_type",  # ← AÑADE ESTA LÍNEA
    "precio": "price",
    "importe": "price",
    "price": "price",  # ← AÑADE ESTA LÍNEA
    "fecha": "date",
    "date": "date",
}


# ============================================================
#  NORMALIZACIÓN UNIVERSAL DE HEADERS (CORRECCIÓN CLAVE)
# ============================================================

def _normalise_header(h: str) -> str:
    """
    Normaliza encabezados eliminando:
    - BOM (\ufeff)
    - NBSP (\xa0)
    - acentos invisibles
    - espacios duplicados
    - mayúsculas/minúsculas
    """
    if not h:
        return ""

    # Normalización unicode (elimina acentos invisibles)
    h = unicodedata.normalize("NFKD", h)

    # Quitar BOM y espacios NBSP
    h = h.replace("\ufeff", "").replace("\xa0", " ")

    # Trim, minúsculas, collapse spaces
    return " ".join(h.strip().lower().split())

def _validate_required_string(field: str, value) -> str:
    """
    Valida que un campo obligatorio sea un string no vacío.
    Retorna el valor normalizado (strip).
    
    :param field: nombre lógico del campo (para mensajes de error)
    :param value: valor a validar
    :raises ValueError: si el valor es inválido
    """
    if value is None:
        raise ValueError(f"El campo obligatorio '{field}' es None")

    if not isinstance(value, str):
        raise ValueError(
            f"El campo obligatorio '{field}' debe ser texto, "
            f"recibido: {type(value).__name__}"
        )

    normalized = value.strip()

    if not normalized:
        raise ValueError(f"El campo obligatorio '{field}' está vacío")

    return normalized

def _validate_positive_price(field: str, value) -> float:
    """
    Valida que un campo de precio sea numérico y > 0.
    Retorna el valor normalizado como float.

    :param field: nombre lógico del campo (para mensajes de error)
    :param value: valor a validar
    :raises ValueError: si el valor es inválido
    """
    if value is None:
        raise ValueError(f"El campo '{field}' es obligatorio y no puede ser None")

    # Evitar que strings vacíos pasen al float()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            raise ValueError(f"El campo '{field}' está vacío")

    try:
        price = float(value)
    except (TypeError, ValueError):
        raise ValueError(
            f"El campo '{field}' debe ser numérico, recibido: {value!r}"
        )

    if price <= 0:
        raise ValueError(
            f"El campo '{field}' debe ser mayor que 0, recibido: {price}"
        )

    return price


def _build_header_map(headers):
    """
    Construye dict: logical_name -> real_header_name
    según los alias definidos.
    """
    logical_to_real = {}
    for real in headers:
        if real is None:
            continue
        key = _normalise_header(real)
        if key in HEADER_ALIASES:
            logical = HEADER_ALIASES[key]
            if logical not in logical_to_real:
                logical_to_real[logical] = real
    return logical_to_real


def _check_file_size(file_path: str) -> None:
    size = os.path.getsize(file_path)
    if size > MAX_IMPORT_FILE_SIZE_BYTES:
        raise ValueError(
            f"El archivo supera el tamaño máximo permitido "
            f"({MAX_IMPORT_FILE_SIZE_MB} MB)"
        )

def _get_importer_logger() -> logging.Logger:
    global _LOGGER
    if _LOGGER:
        return _LOGGER

    logger = logging.getLogger("fdt.importer")
    logger.setLevel(logging.INFO)

    # Ruta de logs compatible con .exe compilado
    logs_path = BASE_PATH / "logs"
    logs_path.mkdir(parents=True, exist_ok=True)

    log_file = logs_path / "importer.log"

    handler = RotatingFileHandler(
        str(log_file),
        maxBytes=1_000_000,  # 1 MB
        backupCount=3,
        encoding="utf-8",
    )

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s"
    )
    handler.setFormatter(formatter)

    logger.addHandler(handler)
    logger.propagate = False

    _LOGGER = logger
    return logger
# ============================================================
#  PARSEO DE FECHAS ROBUSTO
# ============================================================
def _parse_date(value: str, default_today: str) -> str:
    if value is None or str(value).strip() == "":
        return default_today

    s = str(value).strip()

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat()
        except ValueError:
            pass

    # Excel serial date
    try:
        if isinstance(value, (int, float)):
            base = date(1899, 12, 30)
            dt = base.fromordinal(base.toordinal() + int(value))
            return dt.isoformat()
    except Exception:
        pass

    raise ValueError(f"Fecha con formato desconocido: {value}")


# ============================================================
#  LECTURA DE CSV y EXCEL
# ============================================================
def _read_csv_rows(file_path: str):
    f = open(file_path, "r", encoding="utf-8", errors="replace")

    # Autodetectar delimitador (coma, punto y coma, tabulador, etc.)
    sample = f.read(4096)
    f.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fallback: coma estándar

    reader = csv.DictReader(f, dialect=dialect)
    headers = reader.fieldnames or []

    def row_generator():
        try:
            for row in reader:
                yield row
        finally:
            f.close()

    return headers, row_generator()


def _read_excel_rows(file_path: str):
    load_workbook = _get_openpyxl()
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    
    headers = [str(h) if h is not None else "" for h in headers]
        
    def row_generator():
        try:
            for r in rows_iter:
                row_dict = {}
                for idx, val in enumerate(r):
                    if idx < len(headers):
                        row_dict[headers[idx]] = val
                yield row_dict
        finally:
            wb.close()

    return headers, row_generator()



# ============================================================
#  FUNCIÓN PRINCIPAL: IMPORTAR A SQLITE
# ============================================================

def _begin_transaction(conn) -> None:
    conn.execute("BEGIN")
 
def import_csv_to_db(file_path: str, dry_run: bool = False):
    import time
    start_time = time.perf_counter()
    logger = _get_importer_logger()
    logger.info("Inicio importación: %s", file_path)
    _check_file_size(file_path)

    try:
        ext = os.path.splitext(file_path)[1].lower()

        # Leer según tipo
        if ext == ".csv":
            headers, raw_rows = _read_csv_rows(file_path)
        elif ext in (".xlsx", ".xlsm"):
            headers, raw_rows = _read_excel_rows(file_path)
        else:
            return False, f"Formato no soportado: {ext}. Usa CSV o XLSX."

        if not headers:
            return False, "El archivo no contiene encabezados."

        # Mapeo real -> lógico
        header_map = _build_header_map(headers)

        # Verificar columnas obligatorias
        missing = REQUIRED_LOGICAL_COLS - set(header_map.keys())
        if missing:
            missing_str = ",".join(sorted(missing))
            found_str = ",".join(sorted(header_map.keys()))
            return False, (
                f"Faltan columnas obligatorias: {missing_str}\n"
                f"Columnas detectadas: {found_str}"
            )

        today_str = date.today().isoformat()
        used_default_date = False
        valid_rows = []
        errors = []
        failed_rows = []

        has_date_col = "date" in header_map

        # Procesar fila por fila
        for i, raw in enumerate(raw_rows, start=2):
            try:
                client_name = _validate_required_string(
                    "client_name", raw.get(header_map["client_name"])
                )
                client_id = _validate_required_string(
                    "client_id", raw.get(header_map["client_id"])
                )
                product_type = _validate_required_string(
                    "product_type", raw.get(header_map["product_type"])
                )
                price = _validate_positive_price(
                    "price", raw.get(header_map["price"])
                )

                if has_date_col:
                    date_raw = raw.get(header_map["date"])
                    parsed_date = _parse_date(date_raw, today_str)
                else:
                    parsed_date = today_str
                    used_default_date = True

                valid_rows.append(
                    (client_name, client_id, product_type, price, parsed_date)
                )

            except Exception as e:
                msg = f"Fila {i}: {e}"
                errors.append(msg)
                logger.warning(msg)

                failed_rows.append({
                    "row": i,
                    "error": str(e),
                    **raw
                })

        if not valid_rows:
            return False, "Ninguna fila válida para importar."

        # Insertar en BD
        conn = get_connection()
        try:
            clients_in_file = {}
            for client_name, client_id, *_ in valid_rows:
                if client_id not in clients_in_file:
                    clients_in_file[client_id] = client_name
            
            _begin_transaction(conn)
            cursor = conn.cursor()
            
            # Obtener client_id existentes en la BD (defensivo si está vacío)
            existing_ids = set()
            if clients_in_file:
                cursor.execute(
                    "SELECT id FROM clients WHERE id IN ({})".format(
                        ",".join("?" * len(clients_in_file))
                    ),
                    list(clients_in_file.keys())
                )
                existing_ids = {row[0] for row in cursor.fetchall()}

            # Crear clientes faltantes
            missing_clients = [
                (cid, cname, "-", None, None, None)
                for cid, cname in clients_in_file.items()
                if cid not in existing_ids
            ]

            if missing_clients:
                cursor.executemany(
                    """
                    INSERT INTO clients (id, name, cif, address, contact_person, email)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    missing_clients
                )
                logger.info("Insertados %d clientes nuevos", len(missing_clients))

            # Insertar ventas
            cursor.executemany(
                """
                INSERT INTO sales (client_name, client_id, product_type, price, date)
                VALUES (?, ?, ?, ?, ?)
                """,
                valid_rows,
            )
            logger.info("Insertadas %d ventas", len(valid_rows))

            if dry_run:
                conn.rollback()
                logger.info("DRY RUN: transacción revertida, datos no guardados")
            else:
                conn.commit()
                logger.info("Transacción confirmada exitosamente")

            duration = time.perf_counter() - start_time

            if not dry_run:
                try:
                    cursor.execute(
                        """
                        INSERT INTO import_metadata
                            (imported_at, file_name, total_rows, valid_rows, rejected_rows, duration_seconds)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            datetime.now().isoformat(timespec="seconds"),
                            os.path.basename(file_path),
                            len(valid_rows) + len(errors),
                            len(valid_rows),
                            len(errors),
                            duration,
                        )
                    )
                    conn.commit()
                except Exception:
                    logger.exception("No se pudo guardar metadata de importación")

            logger.info(
                "Duración=%.2fs | filas=%s | filas/s=%.1f",
                duration,
                len(valid_rows),
                len(valid_rows) / duration if duration else 0,
            )
            logger.info(
                "Importación completada | válidas=%s | rechazadas=%s | default_date=%s | dry_run=%s",
                len(valid_rows),
                len(errors),
                used_default_date,
                dry_run,)
            
        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

        # Mensaje final

        if dry_run:
            msg = f"OK {len(valid_rows)} registros válidos analizados."
        else:
            msg = f"OK {len(valid_rows)} registros importados correctamente."
        if failed_rows:
            failed_path = os.path.join(
                os.path.dirname(DB_NAME),
                f"failed_rows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            with open(failed_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f, fieldnames=failed_rows[0].keys()
                )
                writer.writeheader()
                writer.writerows(failed_rows)

            logger.info(
                "Filas inválidas exportadas a %s", failed_path
            )
        if errors:
            msg += f" ⚠ {len(errors)} filas rechazadas."
        if used_default_date:
            msg += " ⚠ No había columna de fecha: se usó la fecha de hoy."
        if dry_run:
            msg += "\nDRY RUN: ningún dato fue guardado en la base de datos."

        return True, msg

    except Exception as e:
        logger.exception("Error crítico durante importación")
        return False, f"Error al procesar el archivo: {e}"


# ============================================================
#  IMPORTADOR DE LEADS (agnóstico a plataforma)
# ============================================================

def _build_leads_header_map(headers):
    """
    Construye dict: logical_name -> real_header_name
    según los aliases de LEADS definidos.
    
    Solo mapea campos definidos en LEADS_HEADER_ALIASES.
    """
    logical_to_real = {}
    for real in headers:
        if real is None:
            continue
        key = _normalise_header(real)
        
        # Buscar en LEADS_HEADER_ALIASES
        for logical_name, aliases_list in LEADS_HEADER_ALIASES.items():
            normalized_aliases = [_normalise_header(a) for a in aliases_list]
            if key in normalized_aliases:
                if logical_name not in logical_to_real:
                    logical_to_real[logical_name] = real
                break
    
    return logical_to_real


def _build_marketing_spend_header_map(headers):
    """
    Construye dict: logical_name -> real_header_name
    según los aliases de MARKETING_SPEND definidos.
    """
    logical_to_real = {}
    for real in headers:
        if real is None:
            continue
        key = _normalise_header(real)

        for logical_name, aliases_list in MARKETING_SPEND_HEADER_ALIASES.items():
            normalized_aliases = [_normalise_header(a) for a in aliases_list]
            if key in normalized_aliases:
                if logical_name not in logical_to_real:
                    logical_to_real[logical_name] = real
                break

    return logical_to_real


def _validate_optional_string(field: str, value) -> str | None:
    """
    Valida un campo opcional (puede ser None o vacío).
    Si es válido, retorna el string normalizado.
    Si es None o vacío, retorna None.
    
    :param field: nombre del campo (para logs)
    :param value: valor a validar
    :return: string normalizado o None
    :raises ValueError: solo si el valor existe pero es de tipo incorrecto
    """
    if value is None or value == "":
        return None
    
    if not isinstance(value, str):
        s = str(value).strip() if value is not None else None
        if not s:
            return None
        return s
    
    normalized = value.strip()
    if not normalized:
        return None
    
    return normalized


def _normalize_source(source: str) -> str:
    """
    Normaliza el campo 'source' de un lead para agrupar correctamente
    leads y gastos provenientes de distintas plataformas de marketing de forma robusta.
    
    REGLAS OBLIGATORIAS DE NORMALIZACIÓN (en orden):
    1. Convertir a minúsculas
    2. Eliminar espacios al inicio y final (strip)
    3. Eliminar variaciones comunes utilizadas por plataformas de ads:
       - " lead ads"
       - " lead form"
       - " ads"
       - "_ads"
       - "-ads"
    4. Buscar en SOURCE_NORMALIZATION
    5. Si existe mapeo → retornar valor normalizado
    6. Si no existe → retornar valor original normalizado y registrar aviso en log
    
    Esta función DEBE reutilizarse en todos los importadores de datos de marketing:
    - import_leads_to_db() ← aplicado
    - import_marketing_spend_to_db() ← futuro
    - import_campaigns_to_db() ← futuro
    
    Ejemplos de normalización:
    - "Facebook Lead Ads" → "facebook" → "meta_ads"
    - "Instagram Lead Form" → "instagram" → "meta_ads"
    - "Google Search Ads" → "google" → "google_ads"
    - "Meta Ads" → "meta" → "meta_ads"
    - "LinkedIn" → "linkedin" → "linkedin_ads"
    - "unknown_platform" → "unknown_platform" (mantiene valor, log warning)
    
    :param source: valor de source del CSV/Excel
    :return: valor normalizado
    """
    if not source:
        return source
    
    # 1. Convertir a minúsculas y eliminar espacios al inicio/final
    source_key = source.lower().strip()
    
    # 2. Eliminar variaciones comunes utilizadas por plataformas de ads
    #    En orden de longitud más larga a más corta para evitar conflictos
    for suffix in [" lead ads", " lead form", " ads", "_ads", "-ads"]:
        if source_key.endswith(suffix):
            source_key = source_key[:-len(suffix)].strip()
    
    # 3. Buscar en el diccionario de normalización
    normalized = SOURCE_NORMALIZATION.get(source_key, source_key)
    
    # 4. Detección de fuentes desconocidas (no están en SOURCE_NORMALIZATION)
    if normalized == source_key and source_key not in SOURCE_NORMALIZATION:
        logger = _get_importer_logger()
        logger.warning(
            f"Unknown marketing source detected during import: '{source}' -> '{source_key}'"
        )
    
    return normalized


# ============================================================
#  PRE-FLIGHT INSPECTOR (NO abre BD, NO modifica datos)
# ============================================================

def inspect_import_file(file_path: str) -> dict:
    """
    Detecta automáticamente el esquema de un archivo CSV/XLSX sin importar datos.

    No abre conexión a la BD ni modifica ningún dato.
    Reutiliza helpers internos del módulo.

    Retorna dict:
        {
            "valid"          : bool,
            "detected_type"  : "sales" | "leads" | "mixed" | "invalid",
            "headers"        : list[str],
            "missing_sales"  : list[str],   # vacío si cumple esquema ventas
            "missing_leads"  : list[str],   # vacío si cumple esquema leads
        }
    """
    _INVALID = {
        "valid": False,
        "detected_type": "invalid",
        "headers": [],
        "missing_sales": [],
        "missing_leads": [],
        "missing_marketing_spend": [],
    }

    try:
        _check_file_size(file_path)

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            headers, row_gen = _read_csv_rows(file_path)
        elif ext in (".xlsx", ".xlsm"):
            headers, row_gen = _read_excel_rows(file_path)
        else:
            return {**_INVALID, "missing_sales": ["formato_no_soportado"], "missing_leads": ["formato_no_soportado"]}

        if not headers:
            # Consumir generador para cerrar el file handle antes de retornar
            try:
                for _ in row_gen:
                    pass
            except Exception:
                pass
            return {**_INVALID, "missing_sales": ["sin_encabezados"], "missing_leads": ["sin_encabezados"]}

        # --- Esquema VENTAS ---
        sales_map     = _build_header_map(headers)
        missing_sales = sorted(REQUIRED_LOGICAL_COLS - set(sales_map.keys()))
        fits_sales    = len(missing_sales) == 0

        # --- Esquema LEADS ---
        leads_map     = _build_leads_header_map(headers)
        missing_leads = sorted(REQUIRED_LEADS_COLS - set(leads_map.keys()))
        fits_leads    = len(missing_leads) == 0

        # --- Esquema MARKETING SPEND ---
        spend_map             = _build_marketing_spend_header_map(headers)
        missing_marketing_spend = sorted(REQUIRED_MARKETING_SPEND_COLS - set(spend_map.keys()))
        fits_marketing_spend  = len(missing_marketing_spend) == 0

        if fits_sales and fits_leads:
            detected_type = "mixed"
        elif fits_sales:
            detected_type = "sales"
        elif fits_leads:
            detected_type = "leads"
        elif fits_marketing_spend:
            detected_type = "marketing_spend"
        else:
            detected_type = "invalid"

        # --- Contar filas y cerrar el file handle consumiendo todo el generador ---
        row_count = 0
        try:
            for _ in row_gen:
                row_count += 1
        except Exception:
            pass

        return {
            "valid":          detected_type != "invalid",
            "detected_type":  detected_type,
            "headers":        list(headers),
            "row_count":      row_count,
            "valid_rows":     row_count,
            "invalid_rows":   0,
            "missing_sales":  missing_sales,
            "missing_leads":  missing_leads,
            "missing_marketing_spend": missing_marketing_spend,
        }

    except ValueError as exc:
        return {**_INVALID, "missing_sales": [str(exc)], "missing_leads": [str(exc)]}


def import_leads_to_db(file_path: str, dry_run: bool = False):
    """
    Importa leads desde archivo CSV o Excel.
    Agnóstico respecto a la plataforma de origen (Meta, Google, CRM, etc.).
    
    Estructura esperada:
    - Campos obligatorios: created_at, source
    - Campos opcionales: external_id, status, meeting_date, client_id
    - Genera id interno (UUID) si no existe
    - Usa INSERT OR IGNORE para ignorar duplicados automáticamente
    
    Retorna: (ok: bool, message: str)
    """
    import time

    start_time = time.perf_counter()
    logger = _get_importer_logger()
    logger.info("Inicio importación de leads: %s", file_path)
    _check_file_size(file_path)

    try:
        ext = os.path.splitext(file_path)[1].lower()

        # Leer según tipo
        if ext == ".csv":
            headers, raw_rows = _read_csv_rows(file_path)
        elif ext in (".xlsx", ".xlsm"):
            headers, raw_rows = _read_excel_rows(file_path)
        else:
            return False, f"Formato no soportado: {ext}. Usa CSV o XLSX."

        if not headers:
            return False, "El archivo no contiene encabezados."

        # Mapeo real -> lógico (solo campos conocidos)
        header_map = _build_leads_header_map(headers)

        # Verificar columnas obligatorias
        missing = REQUIRED_LEADS_COLS - set(header_map.keys())
        if missing:
            missing_str = ",".join(sorted(missing))
            found_str = ",".join(sorted(header_map.keys()))
            return False, (
                f"Faltan columnas obligatorias: {missing_str}\n"
                f"Columnas detectadas: {found_str}"
            )

        valid_rows = []
        errors = []
        failed_rows = []
        duplicates_ignored = 0

        # Procesar fila por fila
        for i, raw in enumerate(raw_rows, start=2):
            try:
                # ============================================================
                # TRAZABILIDAD: Almacenar valor original antes de normalizar
                # ============================================================
                # Se guarda raw_source para logging y depuración si ocurre
                # algún problema. No se modifica la BD por ahora.
                raw_source = raw.get(header_map["source"])
                
                # Campos obligatorios
                created_at_raw = raw.get(header_map["created_at"])
                created_at = _parse_date(created_at_raw, date.today().isoformat())
                
                source = _validate_required_string(
                    "source", raw_source
                )
                
                # ============================================================
                # NORMALIZACIÓN DE SOURCE (FUNCIÓN REUTILIZABLE)
                # ============================================================
                # Se aplica _normalize_source() para unificar valores de
                # diferentes plataformas (Meta, Google, LinkedIn, TikTok, etc.)
                # 
                # Reglas obligatorias de normalización:
                # 1. Convertir a minúsculas
                # 2. Eliminar espacios al inicio y final
                # 3. Eliminar variaciones comunes (" lead ads", " lead form", etc.)
                # 4. Aplicar diccionario de normalización
                # 5. Detectar y registrar fuentes desconocidas
                # 
                # Esta función DEBE usarse en TODOS los importadores que
                # inserten datos en campos "source":
                # - import_leads_to_db() ← aplicado
                # - import_marketing_spend_to_db() ← futuro
                # - import_campaigns_to_db() ← futuro
                # 
                # Garantía de coherencia:
                # leads.source == marketing_spend.source
                # Permite calcular métricas: CPL, CAC, ROI por plataforma
                source = _normalize_source(source)

                # Campos opcionales
                external_id = None
                if "external_id" in header_map:
                    ext_id_raw = raw.get(header_map["external_id"])
                    external_id = _validate_optional_string("external_id", ext_id_raw)

                status = None
                if "status" in header_map:
                    status_raw = raw.get(header_map["status"])
                    status = _validate_optional_string("status", status_raw)

                meeting_date = None
                if "meeting_date" in header_map:
                    meeting_raw = raw.get(header_map["meeting_date"])
                    if meeting_raw:
                        try:
                            meeting_date = _parse_date(meeting_raw, None)
                        except ValueError:
                            # Si no es una fecha válida, ignorar silenciosamente
                            meeting_date = None

                client_id = None
                if "client_id" in header_map:
                    client_id_raw = raw.get(header_map["client_id"])
                    client_id = _validate_optional_string("client_id", client_id_raw)

                # Generar id interno si no existe external_id.
                # ID determinista: evita duplicados al re-importar el mismo archivo.
                if external_id:
                    lead_id = external_id
                else:
                    _key = f"{client_id or ''}_{created_at}_{source}"
                    lead_id = hashlib.sha1(_key.encode()).hexdigest()

                valid_rows.append({
                    "id": lead_id,
                    "external_id": external_id,
                    "created_at": created_at,
                    "source": source,
                    "status": status,
                    "meeting_date": meeting_date,
                    "client_id": client_id,
                })

            except Exception as e:
                msg = f"Fila {i}: {e}"
                errors.append(msg)
                logger.warning(msg)

                failed_rows.append({
                    "row": i,
                    "error": str(e),
                    **raw
                })

        if not valid_rows:
            return False, "Ninguna fila válida para importar."

        # Insertar en BD
        conn = get_connection()
        try:
            _begin_transaction(conn)
            cursor = conn.cursor()

            # Verificar si client_ids existen (solo si se proporcionaron)
            client_ids_in_file = {
                row["client_id"] 
                for row in valid_rows 
                if row["client_id"] is not None
            }
            
            existing_client_ids = set()
            if client_ids_in_file:
                placeholders = ",".join("?" * len(client_ids_in_file))
                cursor.execute(
                    f"SELECT id FROM clients WHERE id IN ({placeholders})",
                    list(client_ids_in_file)
                )
                existing_client_ids = {row[0] for row in cursor.fetchall()}

            # Validar que los client_ids proporcionados existan
            invalid_client_ids = client_ids_in_file - existing_client_ids
            if invalid_client_ids:
                logger.warning(
                    "client_ids no encontrados en BD: %s", invalid_client_ids
                )
                # Se permite insertar leads con client_id no existentes
                # (pueden ser creados posteriorm ente)
                # Cambiar a None si se prefiere ser más estricto:
                # for row in valid_rows:
                #     if row["client_id"] in invalid_client_ids:
                #         row["client_id"] = None

            # Insertar leads (ignora duplicados automáticamente debido al índice único)
            insert_count = 0
            for row in valid_rows:
                cursor.execute(
                    """
                    INSERT OR IGNORE INTO leads
                        (id, external_id, created_at, source, status, meeting_date, client_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row["id"],
                        row["external_id"],
                        row["created_at"],
                        row["source"],
                        row["status"],
                        row["meeting_date"],
                        row["client_id"],
                    ),
                )
                # Contar si se insertó o fue ignorado (duplicado)
                if cursor.rowcount > 0:
                    insert_count += 1
                else:
                    duplicates_ignored += 1

            logger.info(
                "Leads: %d nuevos, %d duplicados ignorados",
                insert_count,
                duplicates_ignored,
            )

            if dry_run:
                conn.rollback()
                logger.info("DRY RUN: transacción revertida, datos no guardados")
            else:
                conn.commit()
                logger.info("Transacción de leads confirmada exitosamente")

            duration = time.perf_counter() - start_time

            if not dry_run:
                try:
                    cursor.execute(
                        """
                        INSERT INTO import_metadata
                            (imported_at, file_name, total_rows, valid_rows, rejected_rows, duration_seconds)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            datetime.now().isoformat(timespec="seconds"),
                            f"leads: {os.path.basename(file_path)}",
                            len(valid_rows) + len(errors),
                            len(valid_rows),
                            len(errors),
                            duration,
                        )
                    )
                    conn.commit()
                except Exception:
                    logger.exception("No se pudo guardar metadata de importación de leads")

            logger.info(
                "Duración=%.2fs | filas=%s | filas/s=%.1f",
                duration,
                len(valid_rows),
                len(valid_rows) / duration if duration else 0,
            )

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

        # Mensaje final
        if dry_run:
            msg = f"OK {len(valid_rows)} leads válidos analizados."
        else:
            msg = f"OK {len(valid_rows)} leads importados correctamente."

        if duplicates_ignored > 0:
            msg += f" ({duplicates_ignored} duplicados ignorados)"
        
        if failed_rows:
            failed_path = os.path.join(
                os.path.dirname(DB_NAME),
                f"failed_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            with open(failed_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f, fieldnames=failed_rows[0].keys()
                )
                writer.writeheader()
                writer.writerows(failed_rows)

            logger.info(
                "Filas inválidas de leads exportadas a %s", failed_path
            )

        if errors:
            msg += f" ⚠ {len(errors)} filas rechazadas."
        if dry_run:
            msg += "\nDRY RUN: ningún dato fue guardado en la base de datos."

        return True, msg

    except Exception as e:
        logger.exception("Error crítico durante importación de leads")
        return False, f"Error al procesar el archivo de leads: {e}"


# ============================================================
#  IMPORTADOR DE MARKETING SPEND
# ============================================================

def import_marketing_spend_to_db(file_path: str, dry_run: bool = False):
    """
    Importa gasto de marketing desde archivo CSV o Excel.

    Estructura esperada:
    - Campos obligatorios: source, amount
    - Campos opcionales: date (si falta, usa fecha de hoy)
    - source se normaliza con _normalize_source() (coherente con leads)

    Retorna: (ok: bool, message: str)
    """
    import time

    start_time = time.perf_counter()
    logger = _get_importer_logger()
    logger.info("Inicio importación de marketing_spend: %s", file_path)
    _check_file_size(file_path)

    try:
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            headers, raw_rows = _read_csv_rows(file_path)
        elif ext in (".xlsx", ".xlsm"):
            headers, raw_rows = _read_excel_rows(file_path)
        else:
            return False, f"Formato no soportado: {ext}. Usa CSV o XLSX."

        if not headers:
            return False, "El archivo no contiene encabezados."

        header_map = _build_marketing_spend_header_map(headers)

        missing = REQUIRED_MARKETING_SPEND_COLS - set(header_map.keys())
        if missing:
            missing_str = ", ".join(sorted(missing))
            found_str = ", ".join(sorted(header_map.keys()))
            return False, (
                f"Faltan columnas obligatorias: {missing_str}\n"
                f"Columnas detectadas: {found_str}"
            )

        valid_rows = []
        errors = []
        failed_rows = []
        today = date.today().isoformat()

        for i, raw in enumerate(raw_rows, start=2):
            try:
                # Source — obligatorio, normalizado
                raw_source = raw.get(header_map["source"])
                source = _validate_required_string("source", raw_source)
                source = _normalize_source(source)

                # Amount — obligatorio, numérico > 0
                raw_amount = raw.get(header_map["amount"])
                amount = _validate_positive_price("amount", raw_amount)

                # Date — opcional, default hoy
                raw_date = None
                if "date" in header_map:
                    raw_date = raw.get(header_map["date"])
                spend_date = _parse_date(raw_date, today)

                valid_rows.append({
                    "source": source,
                    "amount": round(amount, 2),
                    "date": spend_date,
                })

            except Exception as e:
                msg = f"Fila {i}: {e}"
                errors.append(msg)
                logger.warning(msg)
                failed_rows.append({"row": i, "error": str(e), **raw})

        if not valid_rows:
            return False, "Ninguna fila válida para importar."

        # Insertar en BD
        conn = get_connection()
        try:
            _begin_transaction(conn)
            cursor = conn.cursor()

            insert_count = 0
            for row in valid_rows:
                cursor.execute(
                    "INSERT INTO marketing_spend (source, amount, date) VALUES (?, ?, ?)",
                    (row["source"], row["amount"], row["date"]),
                )
                insert_count += 1

            if dry_run:
                conn.rollback()
                logger.info("DRY RUN: transacción revertida, datos no guardados")
            else:
                conn.commit()
                logger.info("Transacción de marketing_spend confirmada exitosamente")

            duration = time.perf_counter() - start_time

            if not dry_run:
                try:
                    cursor.execute(
                        """
                        INSERT INTO import_metadata
                            (imported_at, file_name, total_rows, valid_rows, rejected_rows, duration_seconds)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            datetime.now().isoformat(timespec="seconds"),
                            f"marketing_spend: {os.path.basename(file_path)}",
                            len(valid_rows) + len(errors),
                            len(valid_rows),
                            len(errors),
                            duration,
                        ),
                    )
                    conn.commit()
                except Exception:
                    logger.exception("No se pudo guardar metadata de importación de marketing_spend")

            logger.info(
                "Duración=%.2fs | filas=%s | filas/s=%.1f",
                duration,
                len(valid_rows),
                len(valid_rows) / duration if duration else 0,
            )

        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        # Mensaje final
        if dry_run:
            msg = f"OK {insert_count} registros de inversión analizados."
        else:
            msg = f"OK {insert_count} registros de inversión importados correctamente."

        if failed_rows:
            failed_path = os.path.join(
                os.path.dirname(DB_NAME),
                f"failed_marketing_spend_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            with open(failed_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=failed_rows[0].keys())
                writer.writeheader()
                writer.writerows(failed_rows)
            logger.info("Filas inválidas de marketing_spend exportadas a %s", failed_path)

        if errors:
            msg += f" ⚠ {len(errors)} filas rechazadas."
        if dry_run:
            msg += "\nDRY RUN: ningún dato fue guardado en la base de datos."

        return True, msg

    except Exception as e:
        logger.exception("Error crítico durante importación de marketing_spend")
        return False, f"Error al procesar el archivo de marketing_spend: {e}"