import csv
import os
import sqlite3
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

from db import DB_NAME


# Nombres lógicos obligatorios para guardar en BD
REQUIRED_LOGICAL_COLS = {"client_name", "client_id", "product_type", "price"}

# Posibles alias que pueden aparecer en CSV o Excel
HEADER_ALIASES = {
    "cliente": "client_name",
    "nombre cliente": "client_name",
    "nombre": "client_name",
    "client_name": "client_name",  # ← AÑADE ESTA LÍNEA
    "identificador del cliente": "client_id",
    "cliente id": "client_id",
    "id cliente": "client_id",
    "identificador": "client_id",
    "client_id": "client_id",  # ← AÑADE ESTA LÍNEA
    "tipo de producto": "product_type",
    "producto": "product_type",
    "tipo producto": "product_type",
    "tipo": "product_type",
    "product_type": "product_type",  # ← AÑADE ESTA LÍNEA
    "precio": "price",
    "importe": "price",
    "price": "price",  # ← AÑADE ESTA LÍNEA
    "fecha": "date",
    "date": "date",
}


# ============================================================
#  NORMALIZACIÓN UNIVERSAL DE HEADERS (CORRECCIÓN CLAVE)
# ============================================================
def _normalise_header(h: str) -> str:
    """
    Normaliza encabezados eliminando:
    - BOM (\ufeff)
    - NBSP (\xa0)
    - acentos invisibles
    - espacios duplicados
    - mayúsculas/minúsculas
    """
    if not h:
        return ""

    # Normalización unicode (elimina acentos invisibles)
    h = unicodedata.normalize("NFKD", h)

    # Quitar BOM y espacios NBSP
    h = h.replace("\ufeff", "").replace("\xa0", " ")

    # Trim, minúsculas, collapse spaces
    return " ".join(h.strip().lower().split())


def _build_header_map(headers):
    """
    Construye dict: logical_name -> real_header_name
    según los alias definidos.
    """
    logical_to_real = {}
    for real in headers:
        if real is None:
            continue
        key = _normalise_header(real)
        if key in HEADER_ALIASES:
            logical = HEADER_ALIASES[key]
            if logical not in logical_to_real:
                logical_to_real[logical] = real
    return logical_to_real


# ============================================================
#  PARSEO DE FECHAS ROBUSTO
# ============================================================
def _parse_date(value: str, default_today: str) -> str:
    if value is None or str(value).strip() == "":
        return default_today

    s = str(value).strip()

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat()
        except ValueError:
            pass

    # Excel serial date
    try:
        if isinstance(value, (int, float)):
            base = date(1899, 12, 30)
            dt = base.fromordinal(base.toordinal() + int(value))
            return dt.isoformat()
    except Exception:
        pass

    raise ValueError(f"Fecha con formato desconocido: {value}")


# ============================================================
#  LECTURA DE CSV y EXCEL
# ============================================================
def _read_csv_rows(file_path: str):
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        return headers, list(reader)


def _read_excel_rows(file_path: str):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = []

    for r in rows[1:]:
        row_dict = {}
        for idx, val in enumerate(r):
            if idx < len(headers):
                row_dict[headers[idx]] = val
        data_rows.append(row_dict)

    return headers, data_rows


# ============================================================
#  FUNCIÓN PRINCIPAL: IMPORTAR A SQLITE
# ============================================================
def import_csv_to_db(file_path: str):
    try:
        ext = os.path.splitext(file_path)[1].lower()

        # Leer según tipo
        if ext == ".csv":
            headers, raw_rows = _read_csv_rows(file_path)
        elif ext in (".xlsx", ".xlsm"):
            headers, raw_rows = _read_excel_rows(file_path)
        else:
            return False, f"Formato no soportado: {ext}. Usa CSV o XLSX."

        if not headers:
            return False, "El archivo no contiene encabezados."

        # Mapeo real -> lógico
        header_map = _build_header_map(headers)

        # Verificar columnas obligatorias
        missing = REQUIRED_LOGICAL_COLS - set(header_map.keys())
        if missing:
            return False, (
                f"Faltan columnas obligatorias: {missing}\n"
                f"Encontradas: {header_map.keys()}"
            )

        today_str = date.today().isoformat()
        used_default_date = False
        valid_rows = []
        errors = []

        has_date_col = "date" in header_map

        # Procesar fila por fila
        for i, raw in enumerate(raw_rows, start=2):
            try:
                client_name = str(raw[header_map["client_name"]]).strip()
                client_id = str(raw[header_map["client_id"]]).strip()
                product_type = str(raw[header_map["product_type"]]).strip()
                price_raw = raw[header_map["price"]]
                price = float(price_raw)

                if has_date_col:
                    date_raw = raw.get(header_map["date"])
                    parsed_date = _parse_date(date_raw, today_str)
                else:
                    parsed_date = today_str
                    used_default_date = True

                valid_rows.append(
                    (client_name, client_id, product_type, price, parsed_date)
                )

            except Exception as e:
                errors.append(f"Fila {i}: {e}")

        if not valid_rows:
            return False, "Ninguna fila válida para importar."

        # Insertar en BD
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        cursor.executemany(
            """
            INSERT INTO sales (client_name, client_id, product_type, price, date)
            VALUES (?, ?, ?, ?, ?)
            """,
            valid_rows,
        )
        conn.commit()
        conn.close()

        # Mensaje final
        msg = f"✔ {len(valid_rows)} registros importados correctamente."
        if errors:
            msg += f" ⚠ {len(errors)} filas rechazadas."
        if used_default_date:
            msg += " ⚠ No había columna de fecha: se usó la fecha de hoy."

        return True, msg

    except Exception as e:
        return False, f"Error al procesar el archivo: {e}"


def import_csv_from_content(content: str, filename: str = "upload.csv"):
    """
    Importa datos de contenido CSV/Excel desde string o bytes.
    Usado para WebView/FilePicker donde no tenemos ruta del archivo.
    """
    try:
        import io
        import tempfile
        
        # Crear archivo temporal
        suffix = ".xlsx" if filename.endswith(".xlsx") else ".csv"
        with tempfile.NamedTemporaryFile(mode='w+b', suffix=suffix, delete=False) as tmp:
            if isinstance(content, str):
                tmp.write(content.encode('utf-8'))
            else:
                tmp.write(content)
            tmp_path = tmp.name
        
        try:
            # Usar la función existente
            result = import_csv_to_db(tmp_path)
            return result
        finally:
            # Limpiar archivo temporal
            try:
                os.remove(tmp_path)
            except:
                pass
    
    except Exception as e:
        return False, f"Error al procesar contenido: {e}"
